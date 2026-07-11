"""
Script de migración de PlanillaEscalamientos.xlsx al formato simplificado.

NO modifica el archivo original. Genera PlanillaEscalamientos.nueva.xlsx.

Formato nuevo (3 hojas):
  - ATMS     : ID, NOMBRE, CUSTODIO, SLA            (reemplaza UNIFICADO + SLA)
  - CONTACTOS: CLAVE, TIPO, EMAIL_SEMANA, CC_SEMANA, EMAIL_FINDE, CC_FINDE
               (reemplaza CONTACTOS SEMANA + CONTACTOS FINDE + CONTACTOS_SUC)
  - RCU      : copiado tal cual (necesario para reprocesar RCU del banco)

Uso:
    cd backend
    venv/Scripts/python.exe migracion_planilla.py
"""
import os
import sys
from openpyxl import load_workbook, Workbook


SRC = "PlanillaEscalamientos.xlsx"
DST = "PlanillaEscalamientos.nueva.xlsx"


def normalizar(s):
    """Normaliza string: upper, sin puntos/guiones/espacios."""
    if s is None:
        return ""
    t = str(s).upper().strip()
    for ch in [".", "-", "_", " ", "/"]:
        t = t.replace(ch, "")
    return t


def _val(cell):
    """Devuelve valor de celda como string limpio, o '' si es None/NaN."""
    if cell is None or cell.value is None:
        return ""
    v = str(cell.value).strip()
    if v.lower() in ("nan", "none"):
        return ""
    return v


# ─────────────────────────────────────────────────────────────────────────────
# PASO 0: validar entrada
# ─────────────────────────────────────────────────────────────────────────────
if not os.path.exists(SRC):
    print(f"ERROR: no encuentro {SRC}. Corré este script desde la carpeta backend/.")
    sys.exit(1)

print("=" * 60)
print(f"  MIGRACIÓN DE PLANILLA")
print(f"  Origen : {SRC}")
print(f"  Destino: {DST}")
print("=" * 60)

wb_src = load_workbook(SRC, data_only=True)


# ─────────────────────────────────────────────────────────────────────────────
# PASO 1: leer UNIFICADO + RCU → construir lista de ATMS válidos
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1/4] Leyendo UNIFICADO y RCU...")
ws_unif = wb_src["UNIFICADO"]
ws_rcu = wb_src["RCU"]

# Mapa: ID → ADDRESS2 (nombre) desde RCU para rellenar nombres vacíos
rcu_nombres = {}
for row in ws_rcu.iter_rows(min_row=2, values_only=False):
    id_v = _val(row[0])
    addr2 = _val(row[2]) if len(row) > 2 else ""
    if id_v:
        rcu_nombres[id_v.upper()] = addr2

ids_validos_rcu = set(rcu_nombres.keys())
print(f"      RCU: {len(ids_validos_rcu)} ATMs activos")

# Construir lista ATMS desde UNIFICADO, descartando huérfanos
atms = []  # cada entrada: {id, nombre, custodio, sla}
descartados_huerfanos = 0
descartados_sin_id = 0
sin_nombre_recuperados = 0

for row in ws_unif.iter_rows(min_row=2, values_only=False):
    id_v = _val(row[0])
    if not id_v:
        descartados_sin_id += 1
        continue
    id_up = id_v.upper()

    nombre = _val(row[1]) if len(row) > 1 else ""
    custodio = _val(row[2]) if len(row) > 2 else ""
    sla = _val(row[3]) if len(row) > 3 else ""

    # Si el ATM no está en el RCU actual → huérfano, descartar
    if id_up not in ids_validos_rcu:
        descartados_huerfanos += 1
        continue

    # Si no tiene nombre, intentar recuperarlo del RCU
    if not nombre and id_up in rcu_nombres:
        nombre = rcu_nombres[id_up]
        if nombre:
            sin_nombre_recuperados += 1

    atms.append({
        "id": id_up,
        "nombre": nombre,
        "custodio": custodio,
        "sla": sla
    })

print(f"      UNIFICADO origen: {ws_unif.max_row - 1} filas")
print(f"      ATMs migrados   : {len(atms)}")
print(f"      Descartados huérfanos (ID no está en RCU)  : {descartados_huerfanos}")
print(f"      Descartados sin ID                          : {descartados_sin_id}")
print(f"      Nombres recuperados desde RCU               : {sin_nombre_recuperados}")

# Mapa para luego: ID normalizado → custodio (lo necesitamos para contactos)
id_norm_a_custodio = {normalizar(a["id"]): a["custodio"] for a in atms}
# Set de custodios únicos no-sucursal
custodios_terceros = set()
for a in atms:
    c = a["custodio"]
    c_up = c.upper()
    if c and not ("SUCURSAL" in c_up or c_up.startswith("SUC")):
        custodios_terceros.add(c)


# ─────────────────────────────────────────────────────────────────────────────
# PASO 2: leer CONTACTOS SEMANA + CONTACTOS FINDE + CONTACTOS_SUC
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2/4] Leyendo hojas de contactos...")

# CONTACTOS SEMANA: clave (puede ser ID de ATM o nombre de custodio) → (email, cc)
ws_sem = wb_src["CONTACTOS SEMANA"]
mapa_semana = {}  # clave normalizada → (email, cc, clave_original)
for row in ws_sem.iter_rows(min_row=2, values_only=False):
    clave = _val(row[0])
    email = _val(row[1]) if len(row) > 1 else ""
    cc = _val(row[2]) if len(row) > 2 else ""
    if clave and email:
        mapa_semana[normalizar(clave)] = (email, cc, clave)

# CONTACTOS FINDE: clave (nombre de custodio o "SUCURSAL") → (email, cc)
ws_finde = wb_src["CONTACTOS FINDE"]
mapa_finde = {}
for row in ws_finde.iter_rows(min_row=2, values_only=False):
    clave = _val(row[0])
    email = _val(row[1]) if len(row) > 1 else ""
    cc = _val(row[2]) if len(row) > 2 else ""
    if clave and email:
        mapa_finde[normalizar(clave)] = (email, cc, clave)

# CONTACTOS_SUC: ATM ID → (email, cc) para sucursales
ws_suc = wb_src["CONTACTOS_SUC"]
mapa_suc = {}
for row in ws_suc.iter_rows(min_row=2, values_only=False):
    atm_id = _val(row[0])
    email = _val(row[6]) if len(row) > 6 else ""
    cc = _val(row[7]) if len(row) > 7 else ""
    if atm_id and email:
        mapa_suc[atm_id.upper()] = (email, cc)

print(f"      CONTACTOS SEMANA : {len(mapa_semana)} entradas con email")
print(f"      CONTACTOS FINDE  : {len(mapa_finde)} entradas con email")
print(f"      CONTACTOS_SUC    : {len(mapa_suc)} entradas con email")


# ─────────────────────────────────────────────────────────────────────────────
# PASO 3: armar hoja CONTACTOS unificada
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3/4] Armando hoja CONTACTOS unificada...")

contactos_finales = []  # (clave, tipo, email_sem, cc_sem, email_fin, cc_fin)

# A) Entrada especial: SUCURSAL finde (default para todas las sucursales en finde)
clave_sucursal_norm = normalizar("SUCURSAL")
if clave_sucursal_norm in mapa_finde:
    e_f, c_f, _ = mapa_finde[clave_sucursal_norm]
    contactos_finales.append(("SUCURSAL", "finde-default", "", "", e_f, c_f))

# B) Contactos por custodio (terceros)
custodios_sin_semana = []
custodios_sin_finde = []

for custodio in sorted(custodios_terceros):
    cust_norm = normalizar(custodio)

    # Buscar en CONTACTOS SEMANA: primero por nombre exacto del custodio
    email_sem, cc_sem = "", ""
    if cust_norm in mapa_semana:
        email_sem, cc_sem, _ = mapa_semana[cust_norm]
    else:
        # Si no, buscar un ATM de ese custodio en CONTACTOS SEMANA
        for atm in atms:
            if atm["custodio"] == custodio:
                atm_norm = normalizar(atm["id"])
                if atm_norm in mapa_semana:
                    email_sem, cc_sem, _ = mapa_semana[atm_norm]
                    break

    # Buscar en CONTACTOS FINDE
    email_fin, cc_fin = "", ""
    if cust_norm in mapa_finde:
        email_fin, cc_fin, _ = mapa_finde[cust_norm]

    if not email_sem:
        custodios_sin_semana.append(custodio)
    if not email_fin:
        custodios_sin_finde.append(custodio)

    contactos_finales.append((custodio, "custodio", email_sem, cc_sem, email_fin, cc_fin))

# C) Contactos por sucursal individual
ids_sucursal = [a["id"] for a in atms if "SUCURSAL" in a["custodio"].upper() or a["custodio"].upper().startswith("SUC")]
sucursales_sin_contacto = []

for atm_id in sorted(ids_sucursal):
    if atm_id in mapa_suc:
        e, c = mapa_suc[atm_id]
        contactos_finales.append((atm_id, "sucursal", e, c, "", ""))
    else:
        sucursales_sin_contacto.append(atm_id)

print(f"      Total contactos generados: {len(contactos_finales)}")
print(f"        - default SUCURSAL finde: {1 if clave_sucursal_norm in mapa_finde else 0}")
print(f"        - por custodio (terceros): {len(custodios_terceros)}")
print(f"        - por sucursal individual: {len([c for c in contactos_finales if c[1] == 'sucursal'])}")
print(f"      Custodios sin contacto SEMANA: {len(custodios_sin_semana)}")
if custodios_sin_semana:
    for c in custodios_sin_semana:
        print(f"           - {c}")
print(f"      Custodios sin contacto FINDE : {len(custodios_sin_finde)}")
if custodios_sin_finde:
    for c in custodios_sin_finde:
        print(f"           - {c}")
print(f"      Sucursales sin contacto      : {len(sucursales_sin_contacto)}")
if sucursales_sin_contacto and len(sucursales_sin_contacto) <= 10:
    for s in sucursales_sin_contacto:
        print(f"           - {s}")
elif sucursales_sin_contacto:
    print(f"           (primeras 5: {', '.join(sucursales_sin_contacto[:5])} ...)")


# ─────────────────────────────────────────────────────────────────────────────
# PASO 4: escribir planilla nueva
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n[4/4] Escribiendo {DST}...")

wb_dst = Workbook()
# Borrar la hoja default
wb_dst.remove(wb_dst.active)

# Hoja ATMS
ws_atms = wb_dst.create_sheet("ATMS")
ws_atms.append(["ID", "NOMBRE", "CUSTODIO", "SLA"])
for a in atms:
    ws_atms.append([a["id"], a["nombre"], a["custodio"], a["sla"]])

# Hoja CONTACTOS
ws_cont = wb_dst.create_sheet("CONTACTOS")
ws_cont.append(["CLAVE", "TIPO", "EMAIL_SEMANA", "CC_SEMANA", "EMAIL_FINDE", "CC_FINDE"])
for c in contactos_finales:
    ws_cont.append(list(c))

# Hoja RCU: copia tal cual desde la original (mismas filas, misma estructura)
ws_rcu_dst = wb_dst.create_sheet("RCU")
for row in ws_rcu.iter_rows(values_only=True):
    ws_rcu_dst.append(list(row))

wb_dst.save(DST)
print(f"      OK -> {DST} guardado")
print()
print("=" * 60)
print("  MIGRACIÓN COMPLETADA")
print("  Original (sin tocar): " + SRC)
print("  Nueva               : " + DST)
print("=" * 60)
