import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime


class ExcelHandler:
    """Lee y escribe la planilla en formato simplificado: ATMS + CONTACTOS + RCU."""

    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.data = {
            'contactos': {},        # custodio_norm → [email_sem, cc_sem]
            'contactos_suc': {},    # id_norm → [email_sem, cc_sem]
            'contactos_finde': {},  # custodio_norm o "SUCURSAL" → [email_fin, cc_fin]
            'unificado': {}         # id_norm → {nombre, custodio, sla_marcas}
        }

    def normalizar(self, s):
        if pd.isna(s):
            return ""
        t = str(s).upper().strip()
        for char in [".", "-", "_", " ", "/"]:
            t = t.replace(char, "")
        return t

    def normalizar_custodio(self, region, zone):
        zone = str(zone).strip().upper() if pd.notna(zone) else ''
        if 'Off-prime Brinks' in region: return f'Brinks {zone}'
        if 'Off-prime STE' in region: return 'BHD - STE Metro'
        if region == 'Sucursal-Sucursal': return 'SUCURSAL'
        if 'Sucursal-DriveUP Brinks' in region: return f'Brinks {zone}'
        if 'Sucursal-DriveUP STE' in region: return 'BHD - STE Metro'
        if 'Sucursal-Sucursal STE' in region: return 'BHD - STE Metro'
        if region in ['SUCURSAL', 'BHD - STE Metro', 'Brinks METRO', 'Brinks NORTE', 'Brinks ESTE', 'Brinks SUR']:
            return region
        return region

    @staticmethod
    def _clean(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        if s.lower() in ("nan", "none"):
            return ""
        return s

    # =================================================================
    # CARGA DE DATOS
    # =================================================================

    def cargar_datos(self):
        if not os.path.exists(self.excel_path):
            return False, "Planilla no encontrada"

        try:
            # Resetear data
            self.data = {
                'contactos': {},
                'contactos_suc': {},
                'contactos_finde': {},
                'unificado': {}
            }

            xls = pd.ExcelFile(self.excel_path)

            # ── ATMS → unificado ──
            if "ATMS" in xls.sheet_names:
                df = pd.read_excel(xls, "ATMS")
                for _, row in df.iterrows():
                    if pd.isna(row.iloc[0]):
                        continue
                    id_n = self.normalizar(str(row.iloc[0]))
                    self.data['unificado'][id_n] = {
                        'nombre': self._clean(row.iloc[1]) if len(row) > 1 else "",
                        'custodio': self._clean(row.iloc[2]) if len(row) > 2 else "",
                        'sla_marcas': self._clean(row.iloc[3]) if len(row) > 3 else "",
                    }

            # ── CONTACTOS → distribuir en contactos / contactos_suc / contactos_finde ──
            if "CONTACTOS" in xls.sheet_names:
                df = pd.read_excel(xls, "CONTACTOS")
                for _, row in df.iterrows():
                    if pd.isna(row.iloc[0]):
                        continue
                    clave = self._clean(row.iloc[0])
                    tipo = self._clean(row.iloc[1]) if len(row) > 1 else ""
                    e_sem = self._clean(row.iloc[2]) if len(row) > 2 else ""
                    c_sem = self._clean(row.iloc[3]) if len(row) > 3 else ""
                    e_fin = self._clean(row.iloc[4]) if len(row) > 4 else ""
                    c_fin = self._clean(row.iloc[5]) if len(row) > 5 else ""

                    if not clave:
                        continue
                    clave_norm = self.normalizar(clave)

                    if tipo == 'sucursal':
                        if e_sem:
                            self.data['contactos_suc'][clave_norm] = [e_sem, c_sem]
                    elif tipo == 'custodio':
                        if e_sem:
                            self.data['contactos'][clave_norm] = [e_sem, c_sem]
                        if e_fin:
                            self.data['contactos_finde'][clave_norm] = [e_fin, c_fin]
                    elif tipo == 'finde-default':
                        if e_fin:
                            # Clave normalizada "SUCURSAL"
                            self.data['contactos_finde'][clave_norm] = [e_fin, c_fin]

            return True, "Datos cargados correctamente"
        except Exception as e:
            return False, str(e)

    # =================================================================
    # AGREGAR ATM INDIVIDUAL (form "Agregar manual")
    # =================================================================

    def guardar_atm(self, atm_id, nombre, sla, custodio, email="", cc=""):
        """Agrega o actualiza un ATM en la hoja ATMS.

        Si el custodio es SUCURSAL y hay email, también persiste el contacto
        en la hoja CONTACTOS (tipo='sucursal').
        """
        try:
            wb = load_workbook(self.excel_path)
            if "ATMS" not in wb.sheetnames:
                return False, "Hoja ATMS no encontrada"

            ws = wb["ATMS"]
            id_upper = atm_id.upper()

            # ── 1) ATMS ──
            target_row = None
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(row=r, column=1).value or '').strip().upper() == id_upper:
                    target_row = r
                    break

            if target_row is None:
                target_row = ws.max_row + 1
                ws.cell(row=target_row, column=1, value=id_upper)

            ws.cell(row=target_row, column=2, value=nombre)
            ws.cell(row=target_row, column=3, value=custodio)
            ws.cell(row=target_row, column=4, value=sla)

            # ── 2) CONTACTOS (solo si es sucursal y hay email) ──
            es_suc = "SUCURSAL" in custodio.upper() or custodio.upper().startswith("SUC")
            if es_suc and email and "CONTACTOS" in wb.sheetnames:
                ws_c = wb["CONTACTOS"]
                contact_row = None
                for r in range(2, ws_c.max_row + 1):
                    clave_cell = str(ws_c.cell(row=r, column=1).value or '').strip()
                    tipo_cell = str(ws_c.cell(row=r, column=2).value or '').strip()
                    if self.normalizar(clave_cell) == self.normalizar(id_upper) and tipo_cell == 'sucursal':
                        contact_row = r
                        break
                if contact_row is None:
                    contact_row = ws_c.max_row + 1
                    ws_c.cell(row=contact_row, column=1, value=id_upper)
                    ws_c.cell(row=contact_row, column=2, value='sucursal')
                ws_c.cell(row=contact_row, column=3, value=email)
                ws_c.cell(row=contact_row, column=4, value=cc if cc else None)

            wb.save(self.excel_path)
            return True, "ATM guardado"
        except Exception as e:
            return False, str(e)

    # =================================================================
    # PROCESAMIENTO DE RCU NUEVO (botón "Actualizar planilla")
    # =================================================================

    def procesar_rcu(self, ruta_rcu_nuevo):
        df_nuevo = pd.read_excel(ruta_rcu_nuevo, header=2)
        return self.procesar_rcu_desde_df(df_nuevo)

    def procesar_rcu_desde_df(self, df_nuevo):
        """Procesa el RCU del banco. Actualiza ATMS y la hoja RCU bruta."""
        try:
            wb = load_workbook(self.excel_path)

            # Asegurar hoja ATMS
            if "ATMS" not in wb.sheetnames:
                ws_atms = wb.create_sheet("ATMS")
                for i, h in enumerate(["ID", "NOMBRE", "CUSTODIO", "SLA"], 1):
                    ws_atms.cell(row=1, column=i, value=h)
            else:
                ws_atms = wb["ATMS"]

            # Mapa de ATMs existentes
            ids_existentes = {}
            for r in range(2, ws_atms.max_row + 1):
                v = ws_atms.cell(row=r, column=1).value
                if v:
                    ids_existentes[str(v).strip().upper()] = r

            # Procesar cada fila del RCU nuevo
            actualizados = 0
            nuevos = 0
            ids_nuevo = set()

            for _, row in df_nuevo.iterrows():
                id_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if not id_raw or id_raw.upper() in ["ID", "ID2", "ID3", "ID4"]:
                    continue
                id_upper = id_raw.upper()
                ids_nuevo.add(id_upper)

                address2 = str(row.get('ADDRESS2', '')) if pd.notna(row.get('ADDRESS2', None)) else ""
                region = str(row.get('REGION', '')) if pd.notna(row.get('REGION', None)) else ""
                zone = str(row.get('ZONE', '')) if pd.notna(row.get('ZONE', None)) else ""

                custodio_nuevo = self.normalizar_custodio(region, zone)

                # SLA: conservar el actual si existe; si es nuevo, dejar "Sin SLA"
                if id_upper in ids_existentes:
                    target_row = ids_existentes[id_upper]
                    sla_actual = self._clean(ws_atms.cell(row=target_row, column=4).value)
                    sla_val = sla_actual if sla_actual else "Sin SLA"
                    actualizados += 1
                else:
                    target_row = ws_atms.max_row + 1
                    ws_atms.cell(row=target_row, column=1, value=id_upper)
                    sla_val = "Sin SLA"
                    nuevos += 1
                    ids_existentes[id_upper] = target_row

                ws_atms.cell(row=target_row, column=2, value=address2)
                ws_atms.cell(row=target_row, column=3, value=custodio_nuevo)
                ws_atms.cell(row=target_row, column=4, value=sla_val)

            # Eliminar ATMs huérfanos (no están en el RCU nuevo)
            eliminados = 0
            rows_to_delete = []
            for r in range(2, ws_atms.max_row + 1):
                v = str(ws_atms.cell(row=r, column=1).value or '').strip().upper()
                if v and v not in ids_nuevo:
                    rows_to_delete.append(r)
                    eliminados += 1
            for r in reversed(rows_to_delete):
                ws_atms.delete_rows(r)

            # Actualizar hoja RCU bruta (para que quede sincronizada)
            if "RCU" in wb.sheetnames:
                ws_rcu = wb["RCU"]
                # Borrar contenido viejo (menos header)
                for r in range(ws_rcu.max_row, 1, -1):
                    ws_rcu.delete_rows(r)
                # Volcar el nuevo
                row_idx = 2
                for _, row in df_nuevo.iterrows():
                    id_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                    if not id_raw or id_raw.upper() in ["ID", "ID2", "ID3", "ID4"]:
                        continue
                    for col_idx, val in enumerate(row.values, 1):
                        if pd.notna(val):
                            ws_rcu.cell(row=row_idx, column=col_idx,
                                        value=val if isinstance(val, (int, float)) else str(val))
                    row_idx += 1

            total_planilla = ws_atms.max_row - 1
            wb.save(self.excel_path)
            return True, {
                'actualizados': actualizados,
                'nuevos': nuevos,
                'total_procesados': actualizados + nuevos,
                'total_planilla': total_planilla,
                'eliminados': eliminados,
                'contactos_limpiados': 0,
                'contactos_propagados': 0
            }
        except Exception as e:
            return False, str(e)

    # =================================================================
    # TAB CONTACTOS (lectura y actualización)
    # =================================================================

    def obtener_contactos_custodio(self):
        """Devuelve listas de custodios + sucursales con sus emails desde la hoja CONTACTOS."""
        self.cargar_datos()
        try:
            if not os.path.exists(self.excel_path):
                return {"error": "Planilla no encontrada"}
            wb = load_workbook(self.excel_path, data_only=True)
            if "CONTACTOS" not in wb.sheetnames:
                return {"error": "Hoja CONTACTOS no encontrada"}

            ws = wb["CONTACTOS"]

            id_to_nombre = {k: v.get('nombre', '') for k, v in self.data['unificado'].items()}

            contactos_terceros = []
            sucursales_con_contacto = {}  # id_norm → {email, cc}
            contactos_finde_sucursal = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                clave = self._clean(row[0])
                tipo = self._clean(row[1]) if len(row) > 1 else ""
                e_sem = self._clean(row[2]) if len(row) > 2 else ""
                c_sem = self._clean(row[3]) if len(row) > 3 else ""
                e_fin = self._clean(row[4]) if len(row) > 4 else ""
                c_fin = self._clean(row[5]) if len(row) > 5 else ""

                if not clave:
                    continue

                if tipo == 'finde-default':
                    contactos_finde_sucursal = {"email": e_fin, "cc": c_fin}
                elif tipo == 'custodio':
                    contactos_terceros.append({
                        "custodio": clave,
                        "email": e_sem,
                        "cc": c_sem,
                        "aplica_finde": bool(e_fin),
                        "email_finde": e_fin,
                        "cc_finde": c_fin
                    })
                elif tipo == 'sucursal':
                    id_norm = self.normalizar(clave)
                    sucursales_con_contacto[id_norm] = {"email": e_sem, "cc": c_sem}

            # Listar TODAS las sucursales desde ATMS, no solo las que tienen contacto.
            # Las que no tienen entrada en CONTACTOS aparecen con email vacío para que
            # el usuario las pueda completar desde el tab Contactos.
            contactos_sucursales = []
            for id_norm, info in self.data['unificado'].items():
                custodio = info.get('custodio', '')
                cust_up = custodio.upper()
                if not ('SUCURSAL' in cust_up or cust_up.startswith('SUC')):
                    continue
                ec = sucursales_con_contacto.get(id_norm, {"email": "", "cc": ""})
                contactos_sucursales.append({
                    "id": id_norm,
                    "nombre": info.get('nombre', ''),
                    "email": ec["email"],
                    "cc": ec["cc"]
                })

            return {
                "terceros": contactos_terceros,
                "sucursales": contactos_sucursales,
                "sucursal_finde": contactos_finde_sucursal
            }
        except Exception as e:
            return {"error": str(e)}

    def actualizar_contactos_custodio(self, custodio, email, cc, aplica_finde, tipo,
                                       email_finde="", cc_finde="", solo=""):
        """Actualiza una fila de la hoja CONTACTOS.

        tipo: 'tercero' | 'sucursal' | 'sucursal_finde'
        solo: 'semana' | 'finde' | '' (ambos)
        """
        try:
            wb = load_workbook(self.excel_path)
            if "CONTACTOS" not in wb.sheetnames:
                return {"error": "Hoja CONTACTOS no encontrada"}

            ws = wb["CONTACTOS"]
            cambios = 0

            # Mapear "tipo de interfaz" → "tipo en la hoja"
            if tipo == 'tercero':
                clave_buscar = custodio
                tipo_buscar = 'custodio'
            elif tipo == 'sucursal':
                clave_buscar = custodio  # acá custodio es ATM ID
                tipo_buscar = 'sucursal'
            elif tipo == 'sucursal_finde':
                clave_buscar = "SUCURSAL"
                tipo_buscar = 'finde-default'
            else:
                return {"error": f"Tipo desconocido: {tipo}"}

            clave_buscar_norm = self.normalizar(clave_buscar)

            # Buscar fila existente con la misma clave Y el mismo tipo
            target_row = None
            for r in range(2, ws.max_row + 1):
                clave_cell = str(ws.cell(row=r, column=1).value or '').strip()
                tipo_cell = str(ws.cell(row=r, column=2).value or '').strip()
                if self.normalizar(clave_cell) == clave_buscar_norm and tipo_cell == tipo_buscar:
                    target_row = r
                    break

            if target_row is None:
                target_row = ws.max_row + 1
                ws.cell(row=target_row, column=1, value=clave_buscar)
                ws.cell(row=target_row, column=2, value=tipo_buscar)

            # Para sucursal_finde, los campos email/cc llegan en los parámetros principales,
            # no en email_finde/cc_finde. Lo escribimos en las columnas de FINDE (5, 6).
            if tipo == 'sucursal_finde':
                ws.cell(row=target_row, column=5, value=email if email else None)
                ws.cell(row=target_row, column=6, value=cc if cc else None)
                cambios += 1
            else:
                if not solo or solo == 'semana':
                    ws.cell(row=target_row, column=3, value=email if email else None)
                    ws.cell(row=target_row, column=4, value=cc if cc else None)
                    cambios += 1
                if not solo or solo == 'finde':
                    e_fin = email_finde or email
                    c_fin = cc_finde or cc
                    ws.cell(row=target_row, column=5, value=e_fin if e_fin else None)
                    ws.cell(row=target_row, column=6, value=c_fin if c_fin else None)
                    cambios += 1

            wb.save(self.excel_path)
            return {"status": "success", "cambios": cambios}
        except Exception as e:
            return {"error": str(e)}
